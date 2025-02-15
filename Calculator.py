import openpyxl

# create a new workbook
calc_book = openpyxl.Workbook()

# rename the default sheet
cursheet = calc_book.active
cursheet.title = "Calculator"

cursheet.cell(1,1).value = "First number ==>"
cursheet.cell(2,1).value = "Second number ==>"
cursheet.cell(1,2).value = 13
cursheet.cell(2,2).value = 8

# create the formula
cursheet.cell(3,2).value = "=B1+B2"

# save this workbook
calc_book.save(r"C:\wiseowl\Python\calculator.xlsx")

# close it down
calc_book.close()